# -*- coding: utf-8 -*-
"""Crawl git history and generate the Agent Console change-tracking table.

Usage:
    python docs/generate_change_log.py

Outputs:
    - docs/PRD.md          (table section rewritten between markers)
    - docs/Agent_Console_变更追踪.xlsx

How it works:
    - Walks `git log` on the current branch after BASELINE_SHA, plus any
      EXTRA_SHAS (commits on other lineages worth tracking).
    - Assigns C-XXX work-order IDs in chronological commit order.
    - Infers columns from conventional-commit prefixes, changed file paths,
      and the git author. Applies CURATED overrides (keyed by commit sha) or
      CURATED_BY_SUBJECT where a human has written rich detail.
"""
from __future__ import annotations

import os
import re
import subprocess
import sys
from pathlib import Path

BASELINE_SHA = "64887a4"

# Commits on other lineages (e.g. the pre-existing origin/main line that was
# overwritten by a force-push) that should still appear in the change log.
EXTRA_SHAS = ["9c2dd0e"]

# Commits touching only the PRD HTML reader (build script + generated output)
# are excluded from the change-tracking table by convention.
SKIP_PATHS = {"docs/build_prd_html.py", "PRD.html", "docs/PRD.html"}

REPO = Path(__file__).resolve().parents[1]  # repo root
PRD = REPO / "docs" / "PRD.md"
XLSX = REPO / "docs" / "Agent_Console_变更追踪.xlsx"

REPO_URL = "https://github.com/liwe123/Agent-Togterher"

HEADERS = [
    "改动时间", "ID", "状态", "Git 提交", "作者", "改动类型", "影响范围",
    "改动内容", "前端技术", "后端技术", "是否有数据库", "破坏性变更",
    "验证结果", "备注",
]

# Row layout (0-based) used by helpers.
SHA_COL = 3  # "Git 提交" column index in the row tuple

PREFIX_TYPE = {
    "feat": "Requirement", "feat!": "Requirement", "fix": "BUG",
    "refactor": "Optimization", "optimize": "Optimization", "perf": "Optimization",
    "style": "Optimization", "docs": "Optimization", "chore": "Optimization",
    "test": "Optimization", "build": "Optimization", "ci": "Optimization",
}

# Curated rows: keyed by short commit sha. Fields override git inference.
CURATED = {
    "08ed038": {
        "type": "Optimization",
        "content": "全站优化：消除前端 4 处重复 WebSocket 逻辑与死代码，统一连接/任务工具/常量，增加 ErrorBoundary；后端事件契约规范化、记录模型成本、限制并发",
        "frontend": "提取共用 useWorkspaceSocket；删除死代码(SystemStatus/SoftwareDock/selectConsoleAgents)；新增 ErrorBoundary；常量集中 constants.ts；任务工具去重 task-utils.ts；fetchedRef 防护",
        "backend": "TaskStepEventPayload Schema 替代手写 dict；LiteLLM 响应提取成本；并发控制(max 3, 429)；receipt 标注预留",
        "db": "否", "breaking": "否", "verify": "前端测试 2→28；净减 237 行",
        "notes": "仅前端组件 + 后端工具函数改动，无外部 API 变化",
    },
    "8ab4834": {"type": "Optimization", "content": "README 更新本轮优化详情", "frontend": "README 章节补充", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "测试数 37/28"},
    "5f3b6a1": {"type": "Optimization", "content": "README 全文中文化", "frontend": "README 全量翻译", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "全量翻译"},
    "3a99015": {"type": "Optimization", "content": "README 风格重写，加入架构图与关键决策记录", "frontend": "README 结构重排、徽章、ASCII 架构图", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "资深工程师口吻"},
    "19d4dca": {
        "type": "Requirement",
        "content": "新增「设置 → API Key 管理」：用户可在前端填入/删除各 Provider 密钥，存库优先于环境变量",
        "frontend": "设置页 API Key 管理 UI(密码框+眼睛+保存/删除)；use-settings 扩展",
        "backend": "ProviderCredential 模型；GET/PUT/DELETE /api/provider-keys；Key 解析优先级 DB>env",
        "db": "是(provider_credentials)", "breaking": "是",
        "verify": "后端 37→40", "notes": "Key 永不在列表回传；新增表；PRD: docs/prd/PRD-APIKey管理.md",
    },
    "fb94fb2": {"type": "Requirement", "content": "新增 Windows 一键启动脚本", "frontend": "start.bat / start.ps1", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "Docker Compose 封装；自动复制 .env + 等健康；PRD: docs/prd/PRD-Windows一键启动脚本.md"},
    "f6ac2e8": {"type": "BUG", "content": "修复开发服务器端口无法访问（next.config standalone 配置与 dev 冲突）", "frontend": "output:standalone 改为 NEXT_BUILD_STANDALONE env 按需启用", "backend": "-", "db": "否", "breaking": "否", "verify": "build pass", "notes": "仅生产 Docker 构建启用"},
    "48b3c47": {
        "type": "Requirement",
        "content": "新增自定义模型接入（任意 provider/model + fallback 降级）；修复 API Key 眼睛图标切换失效",
        "frontend": "自定义模型添加/删除 UI +「自定义」徽章；眼睛图标切换修复(undefined 与 React 批处理冲突)",
        "backend": "CustomModelConfig 模型；/api/custom-models；chat_completion 自定义解析；修按 name 查 PK bug",
        "db": "是(custom_model_configs)", "breaking": "是",
        "verify": "后端 37→40", "notes": "新增表；含眼睛 BUG 修复；PRD: docs/prd/PRD-自定义模型接入.md",
    },
    "5b5e5d9": {"type": "Optimization", "content": "README 补充「模型与密钥管理」章节，更新测试数", "frontend": "README 章节补充", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "测试数 40/28"},
    "4396cad": {"type": "BUG", "content": "修复一键启动脚本中文乱码（UTF-8 被 cmd/PowerShell 按 ANSI 解析）", "frontend": "脚本消息改纯 ASCII", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "跨代码页安全"},
    "a138831": {"type": "BUG", "content": "修复设置页自定义模型表单 / API Key 输入框黑底黑字无法阅读", "frontend": "text-foreground 修复", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "-"},
    "c8d1f72": {"type": "BUG", "content": "设置页表单字段文字改纯白，提升可读性", "frontend": "text-white / white/90 / 占位符 white/40", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "-"},
    "c374465": {
        "type": "BUG",
        "content": "修复点击眼睛无法显示已保存 API Key（保存后清空 + Key 不回传，value 恒空）",
        "frontend": "眼睛点击时拉取真实 Key 填充输入框再切换可见性",
        "backend": "新增 GET /api/provider-keys/{provider} 按需返回 Key；get_api_key_value()",
        "db": "否", "breaking": "否", "verify": "后端 42 passed",
        "notes": "显式操作才返回 Key，列表仍不回传",
    },
    "abbb336": {
        "type": "Requirement",
        "content": "API Key 管理收敛为仅 DeepSeek 预设，用户可自行添加任意厂商的 API",
        "frontend": "「添加厂商」表单(任意厂商名+Key)；厂商名 title-case；移除其余预设",
        "backend": "移除 Provider 白名单；/models/providers/status 只显示 deepseek+DB 厂商",
        "db": "否(复用 provider_credentials)", "breaking": "是",
        "verify": "后端 40→42", "notes": "API 行为变化：PUT 接受任意厂商名；PRD: docs/prd/PRD-APIKey管理.md",
    },
    "6289107": {"type": "Optimization", "content": "新增 docs/PRD.md 变更追踪表", "frontend": "表格 + 维护约定", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "建立本表，后续由脚本实时生成"},
    "fc35e0b": {"type": "Optimization", "content": "新增 Excel 变更追踪工作簿", "frontend": "openpyxl 生成脚本", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "表结构含颜色/冻结/筛选"},
    "2ba41fc": {"type": "Optimization", "content": "变更追踪表新增「改动内容」列，明确记录每次改了什么", "frontend": "表结构更新(PRD.md + Excel)", "backend": "-", "db": "否", "breaking": "否", "verify": "-", "notes": "业务描述与前后端技术分离"},
    "9c2dd0e": {
        "type": "Requirement",
        "content": "前端视觉与响应式优化，新增通讯录 /contacts 页面与 Agent 头像组件",
        "frontend": "新增 agent-portrait.tsx、contacts-page.tsx(/contacts 路由)；agent-gallery/status-panel/app-sidebar/chat 组件重构；globals.css 视觉令牌与响应式优化",
        "backend": "-", "db": "否", "breaking": "否",
        "verify": "-", "notes": "1203 插入/471 删除，21 文件；经 C-021 并入当前分支；PRD: docs/prd/PRD-前端视觉与响应式优化.md",
    },
    "b7f7d27": {
        "type": "Optimization",
        "content": "同步优化版项目为新的 Git 基线",
        "frontend": "-", "backend": "-", "db": "否", "breaking": "否", "verify": "-",
        "notes": "将优化后的完整项目作为当前主线基线提交",
    },
    "29deb0e": {
        "type": "Optimization",
        "content": "合并远端历史并保留当前优化版项目状态",
        "frontend": "-", "backend": "-", "db": "否", "breaking": "否", "verify": "-",
        "notes": "采用 ours 策略合并远端历史后推送到 GitHub",
    },
    "6cab7da": {
        "type": "Optimization",
        "content": "同步优化版项目后刷新变更追踪表",
        "frontend": "更新 PRD 与 Excel 变更追踪表", "backend": "生成脚本重跑", "db": "否", "breaking": "否", "verify": "-",
        "notes": "将新基线提交一并纳入改动表",
    },
    "7039e2a": {
        "type": "Optimization",
        "content": "修复 GitHub Actions Python 包工作流以运行后端测试",
        "frontend": "-", "backend": "backend/requirements-dev.txt；backend/tests；.github/workflows/python-package-conda.yml", "db": "否", "breaking": "否", "verify": "backend/tests/test_health.py 通过", "notes": "移除缺失的 environment.yml 依赖，改为 pip 安装并在 backend 目录执行 pytest",
    },
}

# Curated by exact commit subject (so docs/script commits render cleanly even
# before their sha is known). Applied when the sha lookup misses.
CURATED_BY_SUBJECT = {
    "fix: 解决 app/core/__init__.py 与 app/db/session.py 的循环导入问题": {
        "type": "BUG",
        "content": "修复后端容器启动失败缺陷：优化 app/core/__init__.py 的导出内容，解除与 app/db/session.py 的部分循环初始化依赖，确保 Docker 镜像启动与 Uvicorn 服务平滑拉起",
        "frontend": "无变动",
        "backend": "优化 backend/app/core/__init__.py 顶级导出，移除导致循环依赖的急切导入",
        "db": "否",
        "breaking": "否",
        "verify": "pytest 全量 106 passed；python -c 'from app.main import app' 成功导入；Docker backend 正常启动",
        "notes": "修复 Docker Compose backend 容器退出报错",
    },
    "feat: 实现工作流模板引擎（D2），支持多 Agent 编排流水线与一键实例化调度": {
        "type": "Requirement",
        "content": "实现工作流模板引擎（D2）：引入 workflow_templates 数据表与系统预设编排模板；支持多 Agent 协作流水线步骤节点（Node）定义与动态参数占位符（{{var}}）渲染；实现一键实例化生成 Task 任务并自动推入队列；前端开发 /workflows 工作流模板中心，支持模板卡片流、参数表单运行弹窗、节点查看抽屉与侧边栏导航集成",
        "frontend": "新增 /workflows 工作流模板中心（支持系统预设/自定义分类、节点链路查看抽屉、动态入参表单一键运行 Task 并在成功后自动跳转任务详情页、创建自定义工作流 Modal）；AppSidebar 侧边栏新增「工作流」主导航入口",
        "backend": "新增 WorkflowTemplate 数据库模型；新增 endpoints/workflows.py 端点（/workflows, /workflows/{id}/run, /workflows/{id} DELETE）；支持变量替换、Task 实例化调度、系统预设保护与 audit_service 自动审计埋点",
        "db": "是(workflow_templates)",
        "breaking": "否",
        "verify": "pytest 全量 106 passed；前端测试 28 passed；前端 lint/build 0 errors 0 warnings 通过；子 Agent 独立验收通过",
        "notes": "PRD: docs/prd/PRD-工作流模板引擎.md；工单 C-111；子 Agent 独立验收通过",
    },
    "feat: 实现插件注册中心（D1），支持 Manifest 校验与工作区工具挂载": {
        "type": "Requirement",
        "content": "实现插件注册中心（D1）：引入 plugins 与 workspace_plugins 两张数据表；支持通过 JSON Manifest 注册外部工具、校验 Tool Schemas、按工作区挂载与动态开启/停用；前端开发 /settings/plugins 插件市场管理控制台、Manifest 预览抽屉与注册对话框；设置中心导航升级为 5 卡片网格",
        "frontend": "新增 /settings/plugins 插件管理控制台页面（支持搜索、Manifest 检查抽屉、启用/停用 Switch 与注册新插件 Modal）；设置中心首页导航升级为 5 卡片网格",
        "backend": "新增 Plugin 与 WorkspacePlugin 数据库模型；新增 endpoints/plugins.py 端点（/plugins, /toggle, /active-tools）；RBAC 权限拦截与 audit_service 自动记录 plugin.toggle 审计日志",
        "db": "是(plugins, workspace_plugins)",
        "breaking": "否",
        "verify": "pytest 全量 105 passed；前端测试 28 passed；前端 lint/build 0 errors 0 warnings 通过；子 Agent 独立验收通过",
        "notes": "PRD: docs/prd/PRD-插件注册中心.md；工单 C-110；子 Agent 独立验收通过",
    },
    "feat: 实现任务执行回放与单步调试（B2）及工作区配额治理（C2）": {
        "type": "Requirement",
        "content": "实现任务执行回放与单步调试（B2）及工作区配额治理（C2）：引入 quota_configs 表与配额限流服务；实现任务结构化时序执行流回放与单步断点恢复执行接口；前端开发 TaskReplayPlayer 回放播放器并嵌入任务详情页；前端开发 /settings/quota 配额管理台与设置中心 4 卡片导航",
        "frontend": "新增 TaskReplayPlayer 回放交互播放器组件（支持播放/暂停、倍速调节、时间轴 scrub 与 Payload 检查）；任务详情页嵌入回放流；新增 /settings/quota 工作区配额与限流控制台；设置中心首页导航升级为 4 卡片网格",
        "backend": "新增 QuotaConfig 数据库模型与 quota_service 预算及限流治理；新增 endpoints/quota.py 与 endpoints/task_replay.py（/replay 与 /resume-from-step 接口）；支持断点恢复调度与审计日志联动",
        "db": "是(quota_configs)",
        "breaking": "否",
        "verify": "pytest 全量 104 passed；前端测试 28 passed；前端 lint/build 0 errors 0 warnings 通过；子 Agent 独立验收通过",
        "notes": "PRD: docs/prd/PRD-任务执行回放与单步调试.md, docs/prd/PRD-工作区配额与限流治理.md；工单 C-108, C-109；子 Agent 独立验收通过",
    },
    "feat: 实现平台级审计日志（B1）与成本统计面板（C1）": {
        "type": "Requirement",
        "content": "新增平台级审计日志（B1）与成本统计面板（C1）：引入 audit_logs 表与审计服务，在登录、注册、成员变更与任务中自动埋点；实现成本中心多维聚合统计（总览指标/每日趋势/模型分布/Top任务）；前端实现 /settings/audit 审计日志台与 /settings/cost 成本统计大屏",
        "frontend": "新增 /settings/audit 审计操作日志控制台与 /settings/cost 成本中心与 Token 分析大屏；设置中心首页增加审计与成本快捷导航卡片",
        "backend": "新增 AuditLog 模型；新增 audit_service.py 异步记录器；新增 endpoints/audit_logs.py 与 endpoints/cost_stats.py 并挂载路由与 RBAC 权限拦截；auth/members/tasks 自动埋点",
        "db": "是(audit_logs)",
        "breaking": "否",
        "verify": "pytest 全量 102 passed；前端测试 28 passed；前端 lint/build 0 errors 0 warnings 通过；子 Agent 独立验收通过",
        "notes": "PRD: docs/prd/PRD-平台级审计日志.md, docs/prd/PRD-成本统计面板.md；工单 C-105, C-106；子 Agent 独立验收通过",
    },
    "feat: 实现 RBAC 角色权限模型与多租户 Workspace 隔离（A2+A3）": {
        "type": "Requirement",
        "content": "新增 RBAC 角色权限模型与多租户 Workspace 隔离（A2+A3）：引入 workspace_memberships 与 workspace_invitations 表；实现四级角色（owner/admin/member/viewer）与权限矩阵；提供工作区创建、我的工作区、成员列表、邀请码生成与加入等 API；前端实现侧边栏工作区切换器、成员管理台（/settings/members）与用户退出登录",
        "frontend": "新增 useWorkspaces 与 usePermissions hooks；AppSidebar 新增工作区切换器浮层、创建/加入弹窗与退出登录；新增 /settings/members 成员管理台；设置中心增加成员管理入口",
        "backend": "新增 WorkspaceMembership 与 WorkspaceInvitation 模型；新增 core/permissions.py 权限矩阵与拦截依赖；新增 endpoints/workspace_members.py；注册自动建立工作区关系",
        "db": "是(workspace_memberships, workspace_invitations)",
        "breaking": "否",
        "verify": "pytest 全量 100 passed；前端测试 28 passed；前端 lint/build 均通过；子 Agent 独立验收通过",
        "notes": "PRD: docs/prd/PRD-角色权限与多租户隔离.md；工单 C-103；子 Agent 独立验收通过",
    },
    "feat: 实现用户认证系统（A1），引入 JWT、User 模型与前后端路由鉴权": {
        "type": "Requirement",
        "content": "新增用户认证系统（A1）：引入 User 数据模型与 PBKDF2/JWT 认证，提供注册、登录、Token 刷新、登出与当前用户信息接口；支持 API Token 与 JWT 双轨鉴权；WebSocket 握手 JWT 鉴权；前端新增登录页、注册页、客户端路由守卫与 API 客户端 401 自动续期",
        "frontend": "新增 /login 与 /register 登录注册页面；新增 AuthGuard 客户端路由守卫；task-api.ts 支持 Bearer Token 自动注入与 401 无感刷新；WebSocket 连接携带 JWT",
        "backend": "新增 User 数据库模型；新增 auth.py 模块(JWT 签发/校验与密码哈希)；新增 /api/v1/auth 路由(register/login/refresh/logout/me)；main.py 中间件双轨鉴权",
        "db": "是(users)",
        "breaking": "否（保持与现有 API Token 向后兼容）",
        "verify": "前后端冒烟测试通过；后端全量 99 passed；前端测试 28 passed；前端 lint/build 均通过",
        "notes": "PRD: docs/prd/PRD-用户认证系统.md；工单 C-101；子 Agent 独立验收通过",
    },
    "feat: 增加持久化任务队列与独立 Worker": {
        "type": "Requirement",
        "content": "新增持久化任务队列与独立 Worker，将消息接入和任务执行解耦，并提供优先级、执行租约、失败重试、超时回收、死信和并发控制能力",
        "frontend": "-",
        "backend": "新增 TaskQueueItem、TaskService 与独立 Worker；MessageHub 统一入队；支持 inline/worker 两种执行模式",
        "db": "是(task_queue_items)",
        "breaking": "否（默认保持 inline 模式兼容）",
        "verify": "A/B 执行模式对比通过；后端 59 tests passed；启动模式与队列复验 8 passed；git diff --check 通过",
        "notes": "PRD: docs/prd/PRD-Phase2持久化任务队列与独立Worker.md；独立只读验收通过；生成脚本与 PRD.html 已重跑对齐",
    },
    "docs: 建立 Phase 0 架构治理基线": {
        "type": "Optimization",
        "content": "完成项目 Phase 0 架构盘点，统一记录当前模块职责、任务主链路、任务状态机、Trace/Correlation 标识规范、风险清单与演进边界，为后续 PostgreSQL、TaskService、Worker 和事件总线改造提供约束基线",
        "frontend": "-",
        "backend": "docs/架构治理基线.md；基于 MessageHub、AgentOrchestrator、任务租约、启动恢复和 WebSocket 当前实现形成架构基线",
        "db": "否",
        "breaking": "否",
        "verify": "A/B 文档覆盖对比通过（改前无 Phase 0 基线，改后覆盖 6 类交付物）；文档冒烟通过；后端 56 tests passed",
        "notes": "任务类型为 Optimization，无需新增 PRD；独立验收通过；生成脚本与 PRD.html 已重跑对齐",
    },
    "feat: improve frontend interaction and accessibility": {
        "details": [
            {"type": "Requirement", "content": "Software Dock 增加可操作的连接状态详情", "frontend": "software-dock.tsx：展示软件名称、接入位、在线状态；支持鼠标与键盘操作", "notes": "PRD FR1；依赖同提交"},
            {"type": "Requirement", "content": "Agent 消息支持安全 Markdown 结构化渲染", "frontend": "message-bubble.tsx + react-markdown：标题、列表、行内代码、代码块、安全外链；跳过原始 HTML", "notes": "PRD FR2；package.json/package-lock.json 新增 react-markdown"},
            {"type": "Requirement", "content": "移动端聊天增加快捷指令与 Agent 选择面板", "frontend": "chat-composer.tsx：移动快捷操作入口、模板列表、动态 Agent 列表", "notes": "PRD FR3"},
            {"type": "Requirement", "content": "Agent 详情弹层补齐完整交互能力", "frontend": "agent-gallery.tsx：Dialog 语义、Escape/遮罩关闭、背景滚动锁定、焦点恢复", "notes": "PRD FR4"},
            {"type": "BUG", "content": "修复移动端浅色变量与根节点深色主题冲突", "frontend": "globals.css：移除按屏幕宽度覆盖主题色与 color-scheme 的规则", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复通讯录搜索后 Agent 服务分组错乱", "frontend": "contacts-page.tsx：先按原始服务归属分组，再分别执行搜索过滤", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复通讯录在线人数与连接状态不一致", "frontend": "contacts-page.tsx：结合 WebSocket 连接状态与 Agent 状态计算在线人数", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复聊天提及列表依赖固定中文姓名", "frontend": "chat-composer.tsx：基于接口 Agent 数据和角色优先级生成提及列表", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复客户端 mention 查询参数变化后输入框不更新", "frontend": "chat-composer.tsx：监听查询参数并去重追加提及，保留用户已有输入", "notes": "无需 PRD"},
            {"type": "Optimization", "content": "设置页表单统一使用全局语义颜色", "frontend": "settings-page.tsx：输入框改用 background/foreground/input/muted 设计令牌", "notes": "无需 PRD"},
            {"type": "Optimization", "content": "Agent 卡片扩大为完整语义化交互区域", "frontend": "agent-gallery.tsx：头像、姓名、状态和角色合并为可聚焦按钮", "notes": "无需 PRD"},
            {"type": "Optimization", "content": "完善通讯录搜索范围、无结果状态与清空操作", "frontend": "contacts-page.tsx：支持姓名/角色/职责，Escape 清空，无结果提示与清空按钮", "notes": "无需 PRD"},
            {"type": "Optimization", "content": "提升 Agent 角色辅助文字可读性", "frontend": "agent-gallery.tsx：角色文字由 10px 调整为 11px", "notes": "无需 PRD"},
            {"type": "Optimization", "content": "新增前端交互体验完善 PRD", "frontend": "docs/prd/PRD-前端交互体验完善.md", "notes": "登记 Requirement 的目标、用户故事、FR 与验收标准"},
        ],
        "backend": "-", "db": "否", "breaking": "否",
        "verify": "A/B 代码与视觉对比通过；375/桌面冒烟通过；lint + 28 tests + build 通过",
    },
    "feat: harden backend access and task isolation": {
        "details": [
            {"type": "Requirement", "content": "新增可选 API Token 访问控制", "backend": "APP_API_TOKEN；REST 支持 Bearer/X-API-Key；WebSocket 支持请求头或 token 查询参数", "notes": "PRD: docs/prd/PRD-后端访问控制与工作区隔离.md；FR1-FR3"},
            {"type": "Requirement", "content": "Agent 工具继承可信工作区执行上下文", "backend": "Orchestrator 注入 workspace_id；query_tasks/get_agents 强制按当前工作区过滤并忽略模型伪造范围", "notes": "PRD FR4"},
            {"type": "Requirement", "content": "Provider Key 接口改为仅返回掩码元数据", "frontend": "设置页查看已保存凭证时仅显示掩码，并禁止将掩码作为新密钥保存", "backend": "ProviderKeyValue 返回 configured/masked_key/source，不再返回 api_key 原文", "notes": "PRD FR5"},
            {"type": "BUG", "content": "修复客户端可伪造 Agent/System 消息", "backend": "公开消息接口仅接受 user + normal，拒绝 sender_id、Agent、System 与其他消息类型", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复客户端可篡改任务执行状态与租约", "backend": "TaskCreate/TaskUpdate 移除 status/result/execution_token/expires_at 等服务端字段", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复工作区活跃任务配额漏算 Pending 任务", "backend": "MessageHub 配额统计改为 Pending + Running，达到上限返回 429", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复重启恢复任务绕过工作区并发额度", "backend": "recover_unfinished_tasks 按 Workspace 计算可用额度并限量调度", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复长任务固定租约过期后可能被重复执行", "backend": "Orchestrator 启动后台 lease heartbeat，按 execution_token 条件定期续租并在结束时取消", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复模型测试接口无法识别数据库自定义模型", "backend": "models/test 同时传入 DB API Keys 与 DB custom_models", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复模型列表忽略数据库 Provider Key 状态", "backend": "models 列表统一以 DB 优先、环境变量兜底判断 configured", "notes": "无需 PRD"},
            {"type": "BUG", "content": "修复删除仍被引用的自定义模型导致悬空配置", "backend": "删除前检查 Agent.model_name 和其他模型 fallback_model，存在引用返回 409", "notes": "无需 PRD"},
            {"type": "Optimization", "content": "补充后端访问控制与工作区隔离回归测试", "backend": "新增 API Token、消息伪造、任务字段保护、凭证脱敏、工具跨工作区隔离测试；更新既有 API 契约测试", "notes": "无需 PRD"},
            {"type": "Optimization", "content": "新增后端访问控制与工作区隔离 PRD", "backend": "docs/prd/PRD-后端访问控制与工作区隔离.md", "notes": "覆盖目标、FR、非功能需求与验收标准"},
        ],
        "frontend": "-", "db": "否", "breaking": "是（消息、任务更新、Provider Key GET 响应契约收紧）",
        "verify": "A/B 代码对比通过；API 冒烟通过；56 tests + compileall + pip check 通过",
    },
    "docs: auto-generate change log from git history": {
        "type": "Optimization", "content": "变更追踪表改为从 git history 自动生成",
        "frontend": "generate_change_log.py（爬取 git log + 自动推断列 + CURATED 人工覆盖 + 生成 PRD/Excel）",
        "backend": "-", "db": "否", "breaking": "否", "verify": "-",
        "notes": "新提交自动生成行；已知提交按 sha 覆盖",
    },
    "docs: add visual-aesthetics commit C-005 to change log": {
        "type": "Optimization", "content": "变更追踪表收录独立主线的视觉重构提交（C-005）",
        "frontend": "generate_change_log.py 支持 EXTRA_SHAS + 按 subject 覆盖",
        "backend": "-", "db": "否", "breaking": "否", "verify": "-",
        "notes": "9c2dd0e 强推覆盖后归位",
    },
    "merge: A/B test visual-aesthetics commit 9c2dd0e": {
        "type": "Requirement",
        "content": "合并视觉重构提交 9c2dd0e（A/B 测试通过）：新增 /contacts 通讯录页、agent-portrait 头像组件、恢复 software-dock；globals.css 视觉与响应式优化；设置页与 API Key/自定义模型功能共存",
        "frontend": "合并 settings-page.tsx（保留 API Key 管理 + 自定义模型功能并采用视觉样式）；新增 contacts 路由、agent-portrait.tsx、software-dock.tsx 恢复；agent-gallery/status-panel/app-sidebar/chat 视觉重构",
        "backend": "-", "db": "否", "breaking": "否",
        "verify": "lint/test/build/pytest 全过(42)",
        "notes": "仅 settings-page.tsx 1 处文本冲突手工合并；PRD: docs/prd/PRD-前端视觉与响应式优化.md",
    },
    "docs: 为 5 个历史需求补充详细 PRD 并登记到变更追踪表": {
        "type": "Optimization",
        "content": "为 5 个历史需求（Agent Console MVP / API Key 管理 / 自定义模型 / 前端视觉与响应式优化 / Windows 一键启动）补充详细 PRD 文档到 docs/prd/，并在追踪表对应行备注 PRD 链接",
        "frontend": "新增 docs/prd/PRD-AgentConsoleMVP.md、PRD-APIKey管理.md、PRD-自定义模型接入.md、PRD-前端视觉与响应式优化.md、PRD-Windows一键启动脚本.md",
        "backend": "generate_change_log.py CURATED 备注追加 PRD 链接",
        "db": "否", "breaking": "否", "verify": "-",
        "notes": "docs/PRD.md 增补「PRD 文档索引」；MVP 为基线前需求，以其 PRD 文档登记",
    },
    "feat: add agent tool-calling (function calling) capability": {
        "type": "Requirement",
        "content": "为 Agent 增加工具调用（Function Calling）能力：chat_completion 支持 tools + tool_calls；新增工具注册表（calculate/query_tasks/get_agents/get_system_status）；orchestrator 工具循环（最大 5 轮）并持久化为 TaskStep；单 Agent 与 Worker 阶段启用",
        "frontend": "task-format.ts stepLabel 增加「工具调用」映射",
        "backend": "litellm_service 支持 tools/tool_calls；新增 services/tools.py 注册表+安全计算；orchestrator 工具循环+TaskStep 持久化；config 新增 agent_tools_enabled",
        "db": "否", "breaking": "否",
        "verify": "pytest 42→54；前端 28/build pass",
        "notes": "详见 docs/prd/PRD-工具调用能力.md；不新增 WS 事件、不改表结构",
    },
    "add task-level context continuity tracing": {
        "type": "Requirement",
        "content": "单任务上下文连续性保障（上游抽象设计）：任务级上下文构建器 + 模型调用前结构化上下文回灌 + 工具结果/失败信息稳定继承 + 多 Agent 阶段共享 + 失败重试可读历史 + 超长摘要裁剪",
        "frontend": "任务详情页「任务上下文」模块（当前阶段/摘要/工具链/失败与恢复记录）",
        "backend": "orchestrator.py 各阶段注入 build_context_message；execution_trace.py 上下文构建与摘要裁剪；schemas/task.py 上下文/轨迹字段",
        "db": "否（复用 Task/TaskStep/ModelCall 拼装，未新增表）", "breaking": "是",
        "verify": "后端 build/测试通过；前端 build pass",
        "notes": "详见 docs/prd/PRD-单任务上下文连续性与执行过程可视化.md（2026-08-10 由两份 PRD 合并而来）；与 8e22c25 为同一能力的抽象层与实现层",
    },
    "add task execution trace context": {
        "type": "Requirement",
        "content": "任务执行过程可视化与工具调用追踪：新增执行轨迹层，模型每次继续执行前回灌结构化上下文（任务摘要/当前阶段/已完成步骤/工具结果/失败原因）；工具调用形成显式可回放链路；上下文过长时摘要裁剪；失败/重试可读取历史轨迹续跑",
        "frontend": "task-detail-page.tsx 新增 ExecutionTracePanel（轨迹摘要卡 + 执行轨迹时间线渲染）；types/task.ts 定义 TaskTraceEvent",
        "backend": "新增 core/execution_trace.py（TraceArtifact/build_context_message/build_execution_trace/build_trace_artifact，含上下文构建+工具链路提取+两级摘要裁剪+脱敏）；orchestrator.py 模型调用前注入 build_context_message、阶段推进写入轨迹；schemas/task.py 增 TaskTraceRead/TaskDetailRead 轨迹字段；tasks.py _task_detail 实时组装 trace；前端 HTTP 轮询实现实时刷新",
        "db": "否（复用 Task/TaskStep/ModelCall 拼装，未新增表）", "breaking": "是",
        "verify": "后端 build/测试通过；前端 build pass；任务详情页轨迹视图与工具链路可正常展示",
        "notes": "详见 docs/prd/PRD-单任务上下文连续性与执行过程可视化.md（2026-08-10 由两份 PRD 合并而来）；FR8 WebSocket 推送、FR9 双层视图暂未落地（P1 级），AC6 由 HTTP 轮询达成",
    },
    "docs: merge PRD-单任务上下文连续性保障 and PRD-任务执行过程可视化 into unified PRD": {
        "type": "Optimization",
        "content": "合并两份重叠 PRD（单任务上下文连续性保障 70e53ee/C-075 与 任务执行过程可视化与工具调用追踪 8e22c25/C-076）为统一版 PRD-单任务上下文连续性与执行过程可视化.md：背景缺口合并为 4 项、核心概念整合 5 个、FR 统一为 10 条、数据模型决策与实施状态（FR1-FR7/FR10 已落地，FR8/FR9 未落地）写入文档",
        "frontend": "docs/prd/ 目录 10→9 份（删除 2 份旧 PRD，新增 1 份合并版）；docs/PRD.md 索引同步 10→9 行",
        "backend": "generate_change_log.py：补登记 70e53ee（C-075 升级为完整 Requirement）、C-075/C-076 备注指向合并版、is_excluded 改为 all() 语义（仅纯生成物提交跳过，合并提交保留）",
        "db": "否", "breaking": "否", "verify": "重跑生成脚本 77 行；Excel 无乱码；HTML 阅读器 9 PRDs",
        "notes": "合并版含 §8.4 决策记录与 §16 实施状态章节；两份旧 PRD 可从 git 历史找回",
    },
    # 重跑生成脚本的提交（regen commit）主题固定登记，避免自身行回退为英文 git 主题
    "docs: make change-log 改动内容 Chinese and regenerate": {
        "type": "Optimization",
        "content": "文档：将变更追踪表全部改动内容中文化，消除英文 git 主题 fallback；重跑生成脚本同步 PRD/xlsx",
        "frontend": "-",
        "backend": "generate_change_log.py 新增 _CONTENT_FIXES_BY_SHA（21 行 sha→中文内容）",
        "db": "否", "breaking": "否",
        "verify": "重跑后改动内容列零英文",
        "notes": "守则① 改动内容中文化；regen 提交自身行由本 subject 登记保持中文",
    },
}

# === 35 行「改动类型」耐久修正（原 Docs → Requirement / BUG / Optimization）===
# 仅覆盖 type 字段：新增条目由 git 推断其余列，已有条目保留原 content/tech 等。
# bug 类型沿用脚本配色字面量 "BUG"（type_fill 无 "Bug"），保持单元格配色一致。
_TYPE_FIXES_BY_SHA = {
    # --- Optimization ---
    "8ab4834": "Optimization",  # C-002
    "5f3b6a1": "Optimization",  # C-003
    "3a99015": "Optimization",  # C-004
    "5b5e5d9": "Optimization",  # C-010
    "6289107": "Optimization",  # C-016
    "fc35e0b": "Optimization",  # C-017
    "2ba41fc": "Optimization",  # C-018
    "ffe1535": "Optimization",  # C-022
    "d64090a": "Optimization",  # C-023
    "08b8e10": "Optimization",  # C-024
    "e4c9e20": "Optimization",  # C-026
    "197f6e3": "Optimization",  # C-028
    "2313960": "Optimization",  # C-029
    "caaafe1": "Optimization",  # C-030
    "1fd9aba": "Optimization",  # C-031
    "42dcda5": "Optimization",  # C-032
    "9d4a57e": "Optimization",  # C-033
    "b7f7d27": "Optimization",  # C-034
    "29deb0e": "Optimization",  # C-035
    "6cab7da": "Optimization",  # C-036
    "2d1e659": "Optimization",  # C-037
    "0e0a15f": "Optimization",  # C-039
    "39dae3c": "Optimization",  # C-056
    "6816c87": "Optimization",  # C-057
    "555c064": "Optimization",  # C-071
    "6622d08": "Optimization",  # C-072
    "51bf567": "Optimization",  # C-074
    # --- BUG（沿用脚本配色字面量 BUG，非 Bug）---
    "7039e2a": "BUG",           # C-038
    "352503e": "BUG",           # C-073
}
for _sha, _t in _TYPE_FIXES_BY_SHA.items():
    CURATED.setdefault(_sha, {})["type"] = _t

# === 改动内容中文化（覆盖无 CURATED 中文内容、fallback 到英文 git subject 的行）===
# 守则①要求「改动内容中文」；生成器对无 CURATED 的行会把英文提交主题直接当内容，
# 这里按 sha 补中文，不改类型（类型由 type_from_subject / _TYPE_FIXES 推断）。
_CONTENT_FIXES_BY_SHA = {
    "ffe1535": "将视觉重构提交 C-021（A/B 测试通过）登记进变更追踪表",
    "d64090a": "变更追踪表新增「Git 提交」列，记录每次改动对应的提交哈希",
    "08b8e10": "变更追踪表扩展为专业 14 列布局（含状态/类型/影响范围/前后端技术/数据库/破坏性变更等）",
    "e4c9e20": "将工具调用能力需求 C-025 登记进变更追踪表",
    "caaafe1": "新增 GitHub Actions 工作流（Python 包 + Conda 环境）用于后端测试",
    "2d1e659": "将近期同步相关的变更表条目改为中文描述",
    "0e0a15f": "将 GitHub Actions 工作流修复登记进变更追踪表",
    "c190d7d": "修复 Provider 查询与任务恢复逻辑的健壮性问题",
    "c36bf54": "更新数据库 schema 测试以覆盖任务租约字段",
    "39dae3c": "将前端交互体验改动登记进变更追踪表",
    "6816c87": "将每一项前端改动逐条拆登记进变更追踪表",
    "555c064": "将后端访问控制加固改动登记进变更追踪表",
    "6622d08": "整理合并版 PRD 文档（统一多份需求文档结构）",
    "51bf567": "新增架构图与任务流转图到 PRD 文档",
    "0c46dd7": "修复部署配置：在 docker-compose/.env 透传 APP_API_TOKEN 与 AGENT_TOOLS_ENABLED，使容器内鉴权可用",
    "19e6eae": "优化：在 .gitignore 排除 .workbuddy/ 与备份 xlsx，机械落实守则④排除项",
    "5121ce1": "文档：修正 README 两处不一致（SoftwareDock 保留面板、/contacts 已实现），新增代码审查报告",
    "55a8b1c": "治理优化：从变更表生成器与 PRD 中移除非法的第 4 枚举 Docs，仅保留 Requirement/Optimization/BUG",
    "69095e2": "修复脚本：修正 start.ps1 健康检查等待循环（until($?) 首轮恒真导致死循环）",
    "c0d1cd2": "安全修复：safe_eval 增加指数 DoS 防护，限制常量指数 ≤64，拦截 9**9**9 类爆炸输入",
    "398904e": "文档：重跑生成脚本，将本轮审查修复登记为 C-082~C-087，变更表无 Docs 残留",
    "a3377e6": "治理优化：清除变更表生成脚本内残留的 Docs 字面量（源侧与输出一致），并重跑生成脚本同步 PRD/xlsx",
}
for _sha, _c in _CONTENT_FIXES_BY_SHA.items():
    CURATED.setdefault(_sha, {})["content"] = _c

# 已在 CURATED_BY_SUBJECT 的 4 条：原地改 type（避免被 sha 条目顶掉）
_TYPE_FIXES_BY_SUBJECT = {
    "docs: auto-generate change log from git history":
        "Optimization",  # C-019
    "docs: add visual-aesthetics commit C-005 to change log":
        "Optimization",  # C-020
    "docs: 为 5 个历史需求补充详细 PRD 并登记到变更追踪表":
        "Requirement",   # C-027
    "docs: merge PRD-单任务上下文连续性保障 and PRD-任务执行过程可视化 into unified PRD":
        "Requirement",   # C-077
}
for _subj, _t in _TYPE_FIXES_BY_SUBJECT.items():
    CURATED_BY_SUBJECT[_subj]["type"] = _t

# details 展开行 2 条：改 details 内对应项（禁止加 CURATED[sha]，否则多行塌缩、ID 错位）
_TYPE_FIXES_IN_DETAILS = [
    ("feat: improve frontend interaction and accessibility",
     "新增前端交互体验完善 PRD", "Requirement"),  # C-055
    ("feat: harden backend access and task isolation",
     "新增后端访问控制与工作区隔离 PRD", "Requirement"),  # C-070
]
for _subj, _content, _t in _TYPE_FIXES_IN_DETAILS:
    for _d in CURATED_BY_SUBJECT[_subj]["details"]:
        if _d["content"] == _content:
            _d["type"] = _t
            break
    else:
        raise KeyError(f"detail not found: {_subj} / {_content}")

MODEL_FILES = ("app/models/", "provider_credentials", "custom_model_configs")


def run_git(args: list[str]) -> str:
    # Force UTF-8 log output so commit subjects/paths stay intact on machines
    # whose git defaults to a non-UTF-8 locale (e.g. Chinese Windows GBK).
    return subprocess.run(
        ["git", "-c", "i18n.logOutputEncoding=UTF-8", "-C", str(REPO), *args],
        capture_output=True, text=True, encoding="utf-8",
        errors="replace", check=True,
    ).stdout


def changed_files(sha: str) -> list[str]:
    out = run_git(["diff-tree", "--no-commit-id", "--name-only", "-r", sha])
    return [l for l in out.splitlines() if l.strip()]


def is_excluded(files: list[str]) -> bool:
    # Merge commits report an empty file list via diff-tree; they must stay
    # in the table (all([]) == True would wrongly drop them).
    if not files:
        return False
    # Only skip a commit when EVERY changed file is a pure artifact
    # (build script / generated HTML). A commit that also touches real
    # docs or code must still appear in the change-tracking table.
    return all(f.replace("\\", "/") in SKIP_PATHS for f in files)


def infer_db(files: list[str]) -> str:
    names = []
    for f in files:
        if "app/models/" in f.replace("\\", "/"):
            base = os.path.basename(f)
            if base in ("__init__.py", "base.py", "enums.py"):
                continue
            names.append(base.replace(".py", ""))
    if names:
        return "是(" + ", ".join(sorted(set(names))) + ")"
    if any(m in " ".join(files) for m in MODEL_FILES):
        return "是"
    return "否"


def infer_breaking(files: list[str]) -> str:
    # API contract or DB schema change => breaking.
    joined = " ".join(files).replace("\\", "/")
    if "backend/app/models/" in joined or "backend/app/schemas/" in joined:
        return "是"
    return "否"


def infer_scope(files: list[str]) -> str:
    parts: list[str] = []
    for f in files:
        f = f.replace("\\", "/")
        if f.startswith("frontend/"):
            scope = "前端"
        elif f.startswith("backend/app/models/"):
            scope = "数据库"
        elif f.startswith("backend/"):
            scope = "后端"
        elif f.startswith("docs/") or f.endswith(".md"):
            scope = "文档"
        elif f in ("start.bat", "start.ps1", "docker-compose.yml", "Dockerfile"):
            scope = "部署"
        elif f.startswith(".impeccable") or f.startswith("config/"):
            scope = "配置"
        else:
            scope = "其他"
        if scope not in parts:
            parts.append(scope)
    return "、".join(parts) if parts else "其他"


def infer_frontend(files: list[str]) -> str:
    hits = [f for f in files if f.startswith("frontend/")]
    if not hits:
        return "-"
    parts = sorted({"/".join(f.split("/")[:3]) for f in hits})
    return "；".join(parts)


def infer_backend(files: list[str]) -> str:
    hits = [f for f in files if f.startswith("backend/")]
    if not hits:
        return "-"
    parts = sorted({"/".join(f.split("/")[:3]) for f in hits})
    return "；".join(parts)


def type_from_subject(subject: str) -> str:
    m = re.match(r"^\s*([a-z]+)(\([^)]*\))?(!)?:", subject)
    if m:
        return PREFIX_TYPE.get(m.group(1) + (m.group(3) or ""), "Optimization")
    return "Optimization"


def _commit_info(sha: str) -> tuple[str, str, str, str]:
    out = run_git([
        "log", "-1", sha, "--pretty=format:%h|%ad|%an|%s",
        "--date=format:%Y-%m-%d %H:%M",
    ])
    h, when, author, subject = out.strip().split("|", 3)
    return h, when, author, subject


def git_rows():
    out = run_git([
        "log", "--reverse", f"{BASELINE_SHA}..HEAD",
        "--pretty=format:%h|%ad|%an|%s", "--date=format:%Y-%m-%d %H:%M",
    ])
    commits: list[tuple[str, str, str, str]] = []
    seen: set[str] = set()
    for line in out.splitlines():
        h, when, author, subject = line.split("|", 3)
        commits.append((h, when, author, subject))
        seen.add(h)
    for sha in EXTRA_SHAS:
        if sha[:7] not in seen:
            commits.append(_commit_info(sha))
            seen.add(sha[:7])
    commits.sort(key=lambda c: (c[1], c[0]))

    rows = []
    next_id = 1
    for sha, when, author, subject in commits:
        files = changed_files(sha)
        if is_excluded(files):
            continue
        curated = CURATED.get(sha) or CURATED_BY_SUBJECT.get(subject)
        details = curated.get("details") if curated else None
        if details:
            for detail in details:
                rows.append((
                    when, f"C-{next_id:03d}", detail.get("status", "已完成"),
                    sha, author, detail["type"], detail.get("scope", infer_scope(files)),
                    detail["content"], detail.get("frontend", "-"),
                    detail.get("backend", curated.get("backend", "-")),
                    detail.get("db", curated.get("db", "否")),
                    detail.get("breaking", curated.get("breaking", "否")),
                    detail.get("verify", curated.get("verify", "-")),
                    detail.get("notes", "-"),
                ))
                next_id += 1
            continue

        if curated:
            ctype = curated.get("type", type_from_subject(subject))
            content = curated.get("content", subject)
            fe = curated.get("frontend", infer_frontend(files))
            be = curated.get("backend", infer_backend(files))
            db = curated.get("db", infer_db(files))
            breaking = curated.get("breaking", infer_breaking(files))
            verify = curated.get("verify", "-")
            notes = curated.get("notes", "-")
        else:
            ctype = type_from_subject(subject)
            content = subject
            fe = infer_frontend(files)
            be = infer_backend(files)
            db = infer_db(files)
            breaking = infer_breaking(files)
            verify = "-"
            notes = "-"
        status = curated.get("status", "已完成") if curated else "已完成"
        scope = infer_scope(files)
        rows.append((when, f"C-{next_id:03d}", status, sha, author, ctype, scope,
                     content, fe, be, db, breaking, verify, notes))
        next_id += 1
    return rows


def commit_link(sha: str) -> str:
    return f"[{sha}]({REPO_URL}/commit/{sha})"


def md_table(rows) -> str:
    head = "| " + " | ".join(HEADERS) + " |\n"
    sep = "|" + "|".join(["---"] * len(HEADERS)) + "|\n"
    body = []
    for r in rows:
        cells = [str(v).replace("|", "/") for v in r]
        # Render the Git 提交 column as a clickable link.
        cells[SHA_COL] = commit_link(r[SHA_COL])
        body.append("| " + " | ".join(cells) + " |")
    return head + sep + "\n".join(body) + "\n"


def write_prd(rows) -> None:
    marker_start = "<!-- CHANGELOG:START -->"
    marker_end = "<!-- CHANGELOG:END -->"
    table = md_table(rows)
    text = PRD.read_text(encoding="utf-8")
    if marker_start in text:
        head, _, _ = text.partition(marker_start)
        _, _, tail = text.partition(marker_end)
        new = f"{head}{marker_start}\n{table}{marker_end}{tail}"
    else:
        new = text + f"\n{marker_start}\n{table}{marker_end}\n"
    PRD.write_text(new, encoding="utf-8")


def write_xlsx(rows) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    type_fill = {
        "Requirement": PatternFill("solid", fgColor="FFFDE7"),
        "Optimization": PatternFill("solid", fgColor="E8F5E9"),
        "BUG": PatternFill("solid", fgColor="FFEBEE"),
        "Docs": PatternFill("solid", fgColor="E3F2FD"),
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "变更追踪"

    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    c = ws["A1"]
    c.value = "Agent Console · 变更追踪表"
    c.font = Font(size=14, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="37474F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 22

    # Excel 1-based column of the Git 提交 field (row tuple index SHA_COL).
    sha_excel_col = SHA_COL + 1
    type_excel_col = HEADERS.index("改动类型") + 1
    center_cols = {1, 2, 3, sha_excel_col, type_excel_col,
                   HEADERS.index("是否有数据库") + 1, HEADERS.index("破坏性变更") + 1}

    for r, row in enumerate(rows, start=3):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(
                vertical="center", wrap_text=True,
                horizontal="center" if col in center_cols else "left",
            )
            cell.border = border
            if col == type_excel_col:
                cell.fill = type_fill.get(val, PatternFill())
            if col == sha_excel_col:
                cell.hyperlink = f"{REPO_URL}/commit/{val}"
                cell.style = "Hyperlink"
            cell.font = Font(size=10)

    widths = [18, 8, 9, 10, 14, 13, 14, 38, 44, 42, 20, 11, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{len(rows)+2}"
    wb.save(XLSX)


def main() -> None:
    rows = git_rows()
    write_prd(rows)
    write_xlsx(rows)
    print(f"generated {len(rows)} rows -> {PRD.name}, {XLSX.name}")


if __name__ == "__main__":
    sys.exit(main())
